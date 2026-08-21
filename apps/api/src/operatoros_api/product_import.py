"""CSV/XLSX product import (spec D.2 Step 3 "Upload a list", plan §0.6/§3):
validate -> preview (with per-row errors) -> commit, duplicate detection on
SKU and name, a downloadable corrected-template CSV for failed rows.

**XLSX library choice, disclosed decision (docs/DECISIONS.md):** `openpyxl`.
Pure Python, no macro/VBA execution (it only ever reads the worksheet cell
grid, never `vbaProject.bin`), widely used and actively maintained, already
a de facto standard for this exact job in the Python ecosystem. `pandas`
would also work but pulls in numpy and a much larger dependency surface for
what is, here, a straightforward row-by-row read -- not worth it for one
import screen. Upload size is capped (`MAX_UPLOAD_BYTES`) before any
parsing happens, so a hostile/oversized file can't be used to exhaust
memory just by being accepted.

Two-phase flow, in-memory rather than a server-side staging table:
`parse_and_validate` runs on upload and returns a `ImportPreviewResult`
(plan §3) plus the FULL validated/annotated row set, which the caller
(`api/routers/products_import.py`) round-trips back on `/commit` rather
than persisting server-side between the two calls -- simpler than a
staging table + token expiry story for a first-cut import screen, at the
cost of the client needing to hold and re-send the (typically small,
onboarding-sized) row set. Documented as the trade-off it is.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

import openpyxl

MAX_UPLOAD_BYTES = 5 * 1024 * 1024  # 5MB -- generous for an onboarding product list, not unbounded.
REQUIRED_COLUMNS = ["name"]
KNOWN_COLUMNS = [
    "name",
    "sku",
    "barcode",
    "category",
    "unit",
    "cost_price",
    "selling_price",
    "opening_quantity",
]


@dataclass
class ParsedRow:
    row_number: int
    raw: dict[str, str]
    name: str | None = None
    sku: str | None = None
    barcode: str | None = None
    category: str | None = None
    unit: str | None = None
    cost_price_minor: int | None = None
    selling_price_minor: int | None = None
    opening_quantity: str | None = None
    errors: list[str] = field(default_factory=list)
    is_duplicate: bool = False

    @property
    def is_valid(self) -> bool:
        return not self.errors


def _to_minor(value: str, field_name: str) -> int | None:
    value = value.strip()
    if not value:
        return None
    try:
        return int((Decimal(value) * 100).to_integral_value())
    except InvalidOperation as exc:
        raise ValueError(f"{field_name} must be a number") from exc


def _read_csv_rows(raw: bytes) -> list[dict[str, str]]:
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _read_xlsx_rows(raw: bytes) -> list[dict[str, str]]:
    # read_only + data_only: never evaluates formulas as anything other than
    # their last-cached value, never touches macros (openpyxl doesn't parse
    # vbaProject.bin at all under normal load_workbook()).
    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    rows = []
    for values in rows_iter:
        if all(v is None for v in values):
            continue
        row = {
            header[i]: ("" if v is None else str(v))
            for i, v in enumerate(values)
            if i < len(header)
        }
        rows.append(row)
    return rows


def parse_rows(filename: str, raw: bytes) -> list[dict[str, str]]:
    if len(raw) > MAX_UPLOAD_BYTES:
        raise ValueError(f"File is larger than the {MAX_UPLOAD_BYTES // (1024 * 1024)}MB limit.")
    lower = filename.lower()
    if lower.endswith(".csv"):
        return _read_csv_rows(raw)
    if lower.endswith(".xlsx"):
        return _read_xlsx_rows(raw)
    raise ValueError("Only .csv and .xlsx files are supported.")


def validate_rows(
    raw_rows: list[dict[str, str]],
    *,
    existing_skus: set[str],
    existing_names: set[str],
) -> list[ParsedRow]:
    parsed: list[ParsedRow] = []
    seen_skus: set[str] = set()
    seen_names: set[str] = set()

    for i, raw in enumerate(raw_rows, start=2):  # row 1 is the header
        row = ParsedRow(row_number=i, raw=raw)
        name = (raw.get("name") or "").strip()
        if not name:
            row.errors.append("name is required")
        row.name = name or None

        row.sku = (raw.get("sku") or "").strip() or None
        row.barcode = (raw.get("barcode") or "").strip() or None
        row.category = (raw.get("category") or "").strip() or None
        row.unit = (raw.get("unit") or "").strip() or None
        row.opening_quantity = (raw.get("opening_quantity") or "").strip() or None

        try:
            row.cost_price_minor = _to_minor(raw.get("cost_price") or "", "cost_price")
        except ValueError as exc:
            row.errors.append(str(exc))
        try:
            row.selling_price_minor = _to_minor(raw.get("selling_price") or "", "selling_price")
        except ValueError as exc:
            row.errors.append(str(exc))
        if row.opening_quantity is not None:
            try:
                Decimal(row.opening_quantity)
            except InvalidOperation:
                row.errors.append("opening_quantity must be a number")

        # Duplicate detection: against existing products AND against
        # earlier rows in this same file (spec: "duplicate detection on SKU
        # and name").
        if row.sku and (row.sku in existing_skus or row.sku in seen_skus):
            row.is_duplicate = True
            row.errors.append(f"duplicate SKU {row.sku!r}")
        if row.name and (row.name in existing_names or row.name in seen_names):
            row.is_duplicate = True
            row.errors.append(f"duplicate name {row.name!r}")
        if row.sku:
            seen_skus.add(row.sku)
        if row.name:
            seen_names.add(row.name)

        parsed.append(row)

    return parsed


_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _csv_formula_safe(value: str) -> str:
    """Neutralizes CSV/spreadsheet-formula injection (OWASP CSV Injection):
    every field in this file is an echo of a business's own upload, so a
    row that failed validation because its `name`/`sku`/etc. started with
    `=`, `+`, `-`, `@`, or a tab/CR could otherwise land as a live formula
    the moment the corrected-template CSV is reopened in Excel/Sheets.
    Prefixing with a single quote forces spreadsheet apps to treat the
    cell as plain text without changing what the business sees."""
    if value and value[0] in _FORMULA_TRIGGER_CHARS:
        return f"'{value}"
    return value


def corrected_template_csv(rows: list[ParsedRow]) -> str:
    """Spec D.2: "Errors are downloadable as a corrected-template CSV."
    Only the rows that failed validation, in the same column shape as the
    original upload, plus an `errors` column so the business can see what
    to fix without cross-referencing the preview separately."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow([*KNOWN_COLUMNS, "errors"])
    for row in rows:
        if row.is_valid:
            continue
        writer.writerow(
            [
                _csv_formula_safe(row.name or ""),
                _csv_formula_safe(row.sku or ""),
                _csv_formula_safe(row.barcode or ""),
                _csv_formula_safe(row.category or ""),
                _csv_formula_safe(row.unit or ""),
                _csv_formula_safe(str(row.raw.get("cost_price", ""))),
                _csv_formula_safe(str(row.raw.get("selling_price", ""))),
                _csv_formula_safe(row.opening_quantity or ""),
                "; ".join(row.errors),
            ]
        )
    return buffer.getvalue()
